import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roadmap Kelulusan"

# Headers
ws['A1'] = "TAHAPAN"
ws['B1'] = "KEGIATAN / PERSYARATAN"
ws['C1'] = "STATUS"
ws['D1'] = "CATATAN"

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in range(1, 5):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

data = [
    # Tahap 1
    ("TAHAP 1: Pra-Tesis & Judul", "Mengikuti Pembekalan Tesis dan lulus Metodologi Penelitian", "Selesai", ""),
    ("", "Membuat Outline Tesis dan disetujui Dosen Konsultan", "Selesai", "Maks 1 bulan"),
    ("", "Mendaftar Tesis melalui formulir online", "Selesai", ""),
    ("", "Mendapatkan Surat Penetapan Dosen Pembimbing", "Selesai", ""),
    
    # Tahap 2
    ("TAHAP 2: Proposal & WS 1", "Melakukan bimbingan penyusunan Proposal (Bab 1, 2, 3)", "Selesai", ""),
    ("", "Mendaftar dan presentasi di Workshop 1", "Selesai", ""),
    ("", "Mendapatkan feedback dan revisi proposal", "Selesai", ""),
    
    # Tahap 3
    ("TAHAP 3: Eksekusi & Jurnal", "Bimbingan Bab 4 & 5 (Minimal 8 kali pertemuan)", "Selesai", "Tercatat di form"),
    ("", "Mengolah data ML (GMM & LLM)", "Selesai", ""),
    ("", "Menyusun dan submit artikel ilmiah", "Selesai", ""),
    ("", "Mendapatkan LoA dari jurnal", "Selesai", "Syarat wajib WS 2"),
    
    # Tahap 4
    ("TAHAP 4: Pemberkasan Ujian", "Laporan Tesis Lengkap (Cover - Lampiran)", "Selesai", ""),
    ("", "Lembar Bimbingan ditandatangani Pembimbing", "Belum", "Selesai draf, butuh TTD"),
    ("", "Bukti LoA Jurnal", "Selesai", ""),
    ("", "Slide Presentasi (PPTX)", "Selesai", "Versi Bergambar & Tersinkron"),
    ("", "Video Presentasi", "Belum", "Gunakan naskah di PPTX"),
    ("", "Poster Publikasi", "Belum", "Gunakan Canva/Photoshop"),
    
    # Tahap 5
    ("TAHAP 5: Workshop 2 (Sidang)", "Daftar di Smart Campus dan terima jadwal", "Belum", "Lakukan setelah Tahap 4 beres"),
    ("", "Presentasi di hadapan Tim Reviewer (INTIS)", "Belum", ""),
    ("", "Revisi Tesis pasca-sidang (jika ada)", "Belum", "Maks 1 minggu"),
    
    # Tahap 6
    ("TAHAP 6: Syarat Wisuda", "Tanda Tangan Pengesahan (Pembimbing, Penguji, Dekan)", "Belum", ""),
    ("", "Pemberkasan/Cek Plagiasi ke Perpustakaan UNISBANK", "Belum", "Maks 2 minggu setelah nilai keluar"),
    ("", "Upload bukti perpus ke Smart Campus", "Belum", ""),
    ("", "LULUS & DAFTAR WISUDA!", "Belum", "🎉")
]

for row_idx, row_data in enumerate(data, start=2):
    ws.cell(row=row_idx, column=1, value=row_data[0])
    ws.cell(row=row_idx, column=2, value=row_data[1])
    ws.cell(row=row_idx, column=3, value=row_data[2])
    ws.cell(row=row_idx, column=4, value=row_data[3])
    
    # Style styling for Status column
    status_cell = ws.cell(row=row_idx, column=3)
    status_cell.alignment = Alignment(horizontal="center")
    if row_data[2] == "Selesai":
        status_cell.font = Font(color="008000", bold=True)
    elif row_data[2] == "Belum":
        status_cell.font = Font(color="FF0000", bold=True)
        
    if row_data[0] != "":
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 65
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40

wb.save("FULL TESIS/Checklist_Roadmap_Kelulusan.xlsx")
print("Checklist_Roadmap_Kelulusan.xlsx created successfully!")
