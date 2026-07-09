import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roadmap Kelulusan"

# Headers
ws['A1'] = "TAHAPAN"
ws['B1'] = "KEGIATAN / PERSYARATAN"
ws['C1'] = "STATUS (Dropdown)"
ws['D1'] = "CATATAN WAKTU/INFO"
ws['E1'] = "DOKUMEN & KETERANGAN SYARAT"

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Create a data validation object with list of choices
dv = DataValidation(type="list", formula1='"Selesai,Proses,Belum"', allow_blank=True)
# Custom error message
dv.error ='Status harus dipilih dari Selesai, Proses, atau Belum'
dv.errorTitle = 'Status Tidak Valid'
dv.prompt = 'Pilih status dari daftar'
dv.promptTitle = 'Pilih Status'
ws.add_data_validation(dv)

data = [
    # Tahap 1
    ("TAHAP 1: Pra-Tesis & Judul", "Mengikuti Pembekalan Tesis dan lulus Metodologi Penelitian", "Selesai", "", "KHS mata kuliah Metodologi (Nilai min. B)"),
    ("", "Membuat Outline Tesis dan disetujui Dosen Konsultan", "Selesai", "Maks 1 bulan", "Form Outline Tesis (Google Form)"),
    ("", "Mendaftar Tesis melalui formulir online", "Selesai", "", "Bukti Submit Form"),
    ("", "Mendapatkan Surat Penetapan Dosen Pembimbing", "Selesai", "", "SK Dosen Pembimbing (Softcopy)"),
    
    # Tahap 2
    ("TAHAP 2: Proposal & WS 1", "Melakukan bimbingan penyusunan Proposal (Bab 1, 2, 3)", "Selesai", "", "Draf Proposal (Word/PDF)"),
    ("", "Mendaftar dan presentasi di Workshop 1", "Selesai", "", "Slide PPT Proposal"),
    ("", "Mendapatkan feedback dan revisi proposal", "Selesai", "", "Form Revisi TTD Reviewer"),
    
    # Tahap 3
    ("TAHAP 3: Eksekusi & Jurnal", "Bimbingan Bab 4 & 5 (Minimal 8 kali pertemuan)", "Selesai", "Tercatat di form", "Buku/Form Lembar Bimbingan Tesis asli"),
    ("", "Mengolah data ML (GMM & LLM)", "Selesai", "", "Kode Python, Data Output, Model"),
    ("", "Menyusun dan submit artikel ilmiah", "Selesai", "", "Draft Artikel Jurnal (Format Jurnal Tujuan)"),
    ("", "Mendapatkan LoA dari jurnal", "Selesai", "Syarat wajib WS 2", "Surat LoA Resmi (PDF) / Screenshot Email"),
    
    # Tahap 4
    ("TAHAP 4: Pemberkasan Ujian", "Laporan Tesis Lengkap (Cover - Lampiran)", "Selesai", "", "File DOCX/PDF Tesis versi lengkap"),
    ("", "Lembar Bimbingan ditandatangani Pembimbing", "Belum", "Selesai draf, butuh TTD", "Dokumen fisik/scan Lembar Bimbingan dengan TTD basah/digital dosen"),
    ("", "Bukti LoA Jurnal", "Selesai", "", "Dokumen LoA"),
    ("", "Slide Presentasi (PPTX)", "Selesai", "Versi Bergambar & Tersinkron", "File .pptx untuk presentasi"),
    ("", "Video Presentasi", "Belum", "Gunakan naskah di PPTX", "File video MP4 (Biasanya diunggah ke YouTube/Google Drive, serahkan link ke SmartCampus)"),
    ("", "Poster Publikasi", "Belum", "Gunakan Canva/Photoshop", "File gambar/PDF poster. (Biasanya diunggah langsung ke form SmartCampus sebagai syarat ujian, tidak di-publish ke publik kecuali diminta khusus oleh prodi)"),
    
    # Tahap 5
    ("TAHAP 5: Workshop 2 (Sidang)", "Daftar di Smart Campus dan terima jadwal", "Belum", "Lakukan setelah Tahap 4 beres", "Form pendaftaran ujian di portal kampus"),
    ("", "Presentasi di hadapan Tim Reviewer (INTIS)", "Belum", "", "Siapkan Mental & Pakaian Formal (Jas/Almamater)"),
    ("", "Revisi Tesis pasca-sidang (jika ada)", "Belum", "Maks 1 minggu", "Lembar Revisi TTD Penguji"),
    
    # Tahap 6
    ("TAHAP 6: Syarat Wisuda", "Tanda Tangan Pengesahan (Pembimbing, Penguji, Dekan)", "Belum", "", "Halaman Pengesahan Tesis (Asli TTD Basah)"),
    ("", "Pemberkasan/Cek Plagiasi ke Perpustakaan UNISBANK", "Belum", "Maks 2 minggu setelah nilai keluar", "Hardcopy (Jilid lux) & Softcopy (CD) Tesis, Hasil Turnitin"),
    ("", "Upload bukti perpus ke Smart Campus", "Belum", "", "Surat Keterangan Bebas Pustaka / Resi Penyerahan"),
    ("", "LULUS & DAFTAR WISUDA!", "Belum", "🎉", "Syarat Yudisium lengkap")
]

for row_idx, row_data in enumerate(data, start=2):
    ws.cell(row=row_idx, column=1, value=row_data[0])
    ws.cell(row=row_idx, column=2, value=row_data[1])
    ws.cell(row=row_idx, column=3, value=row_data[2])
    ws.cell(row=row_idx, column=4, value=row_data[3])
    ws.cell(row=row_idx, column=5, value=row_data[4])
    
    # Add validation to Column C (Status)
    dv.add(ws.cell(row=row_idx, column=3))
    
    # Alignment
    status_cell = ws.cell(row=row_idx, column=3)
    status_cell.alignment = Alignment(horizontal="center")
    
    # Bold for Stage names
    if row_data[0] != "":
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 65

wb.save("FULL TESIS/Checklist_Roadmap_Kelulusan.xlsx")
print("Checklist_Roadmap_Kelulusan.xlsx updated with dropdowns and rules successfully!")
