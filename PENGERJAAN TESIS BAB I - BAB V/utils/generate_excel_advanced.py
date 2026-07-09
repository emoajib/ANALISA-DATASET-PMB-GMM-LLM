import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Roadmap Kelulusan"

# --- PROGRESS DASHBOARD ---
ws.merge_cells('A1:B2')
dashboard_cell = ws['A1']
dashboard_cell.value = "PROGRESS KELULUSAN TESIS ANDA:"
dashboard_cell.font = Font(bold=True, size=16, color="FFFFFF")
dashboard_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
dashboard_cell.alignment = Alignment(horizontal="right", vertical="center")

ws.merge_cells('C1:E2')
pct_cell = ws['C1']
# Formula to calculate percentage of 'Selesai' out of 24 tasks
pct_cell.value = '=COUNTIF(C5:C30, "Selesai")/24'
pct_cell.number_format = '0.0%'
pct_cell.font = Font(bold=True, size=20, color="00B050")
pct_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
pct_cell.alignment = Alignment(horizontal="center", vertical="center")

# --- HEADERS ---
ws['A4'] = "TAHAPAN"
ws['B4'] = "KEGIATAN / PERSYARATAN"
ws['C4'] = "STATUS"
ws['D4'] = "CATATAN WAKTU/INFO"
ws['E4'] = "DOKUMEN & KETERANGAN SYARAT"

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in range(1, 6):
    cell = ws.cell(row=4, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# --- DATA VALIDATION (Dropdown) ---
dv = DataValidation(type="list", formula1='"Selesai,Proses,Belum"', allow_blank=True)
dv.error ='Status harus dipilih dari Selesai, Proses, atau Belum'
dv.errorTitle = 'Status Tidak Valid'
dv.prompt = 'Pilih status dari daftar'
dv.promptTitle = 'Pilih Status'
ws.add_data_validation(dv)

data = [
    ("TAHAP 1: Pra-Tesis & Judul", "Mengikuti Pembekalan Tesis dan lulus Metodologi Penelitian", "Selesai", "", "KHS mata kuliah Metodologi (Nilai min. B)"),
    ("", "Membuat Outline Tesis dan disetujui Dosen Konsultan", "Selesai", "Maks 1 bulan", "Form Outline Tesis (Google Form)"),
    ("", "Mendaftar Tesis melalui formulir online", "Selesai", "", "Bukti Submit Form"),
    ("", "Mendapatkan Surat Penetapan Dosen Pembimbing", "Selesai", "", "SK Dosen Pembimbing (Softcopy)"),
    
    ("TAHAP 2: Proposal & WS 1", "Melakukan bimbingan penyusunan Proposal (Bab 1, 2, 3)", "Selesai", "", "Draf Proposal (Word/PDF)"),
    ("", "Mendaftar dan presentasi di Workshop 1", "Selesai", "", "Slide PPT Proposal"),
    ("", "Mendapatkan feedback dan revisi proposal", "Selesai", "", "Form Revisi TTD Reviewer"),
    
    ("TAHAP 3: Eksekusi & Jurnal", "Bimbingan Bab 4 & 5 (Minimal 8 kali pertemuan)", "Selesai", "Tercatat di form", "Buku/Form Lembar Bimbingan Tesis asli"),
    ("", "Mengolah data ML (GMM & LLM)", "Selesai", "", "Kode Python, Data Output, Model"),
    ("", "Menyusun dan submit artikel ilmiah", "Selesai", "", "Draft Artikel Jurnal (Format Jurnal Tujuan)"),
    ("", "Mendapatkan LoA dari jurnal", "Selesai", "Syarat wajib WS 2", "Surat LoA Resmi (PDF) / Screenshot Email"),
    
    ("TAHAP 4: Pemberkasan Ujian", "Laporan Tesis Lengkap (Cover - Lampiran)", "Selesai", "", "File DOCX/PDF Tesis versi lengkap"),
    ("", "Lembar Bimbingan ditandatangani Pembimbing", "Belum", "Selesai draf, butuh TTD", "Dokumen fisik/scan Lembar Bimbingan dengan TTD basah/digital dosen"),
    ("", "Bukti LoA Jurnal", "Selesai", "", "Dokumen LoA"),
    ("", "Slide Presentasi (PPTX)", "Selesai", "Versi Bergambar & Tersinkron", "File .pptx untuk presentasi"),
    ("", "Video Presentasi", "Belum", "Gunakan naskah di PPTX", "File video MP4 (Biasanya link YouTube/Drive disetor ke SmartCampus)"),
    ("", "Poster Publikasi", "Belum", "Gunakan Canva/Photoshop", "File gambar/PDF poster, diunggah ke form SmartCampus"),
    
    ("TAHAP 5: Workshop 2 (Sidang)", "Daftar di Smart Campus dan terima jadwal", "Belum", "Lakukan setelah Tahap 4 beres", "Form pendaftaran ujian di portal kampus"),
    ("", "Presentasi di hadapan Tim Reviewer (INTIS)", "Belum", "", "Siapkan Mental & Pakaian Formal (Jas/Almamater)"),
    ("", "Revisi Tesis pasca-sidang (jika ada)", "Belum", "Maks 1 minggu", "Lembar Revisi TTD Penguji"),
    
    ("TAHAP 6: Syarat Wisuda", "Tanda Tangan Pengesahan (Pembimbing, Penguji, Dekan)", "Belum", "", "Halaman Pengesahan Tesis (Asli TTD Basah)"),
    ("", "Pemberkasan/Cek Plagiasi ke Perpustakaan UNISBANK", "Belum", "Maks 2 minggu setelah nilai keluar", "Hardcopy (Jilid lux) & Softcopy (CD) Tesis, Hasil Turnitin"),
    ("", "Upload bukti perpus ke Smart Campus", "Belum", "", "Surat Keterangan Bebas Pustaka / Resi Penyerahan"),
    ("", "LULUS & DAFTAR WISUDA!", "Belum", "🎉", "Syarat Yudisium lengkap")
]

start_row = 5
for i, row_data in enumerate(data):
    row_idx = start_row + i
    ws.cell(row=row_idx, column=1, value=row_data[0])
    ws.cell(row=row_idx, column=2, value=row_data[1])
    ws.cell(row=row_idx, column=3, value=row_data[2])
    ws.cell(row=row_idx, column=4, value=row_data[3])
    ws.cell(row=row_idx, column=5, value=row_data[4])
    
    # Validation Dropdown
    dv.add(ws.cell(row=row_idx, column=3))
    
    # Bold stage headers
    if row_data[0] != "":
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        # Give a slight background color to separate stages
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

# --- CONDITIONAL FORMATTING ---
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
yellow_font = Font(color="9C5700", bold=True)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006", bold=True)

# Apply rules to column C (Status)
ws.conditional_formatting.add(f'C5:C30', CellIsRule(operator='equal', formula=['"Selesai"'], fill=green_fill, font=green_font))
ws.conditional_formatting.add(f'C5:C30', CellIsRule(operator='equal', formula=['"Proses"'], fill=yellow_fill, font=yellow_font))
ws.conditional_formatting.add(f'C5:C30', CellIsRule(operator='equal', formula=['"Belum"'], fill=red_fill, font=red_font))


# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 65

wb.save("FULL TESIS/Checklist_Roadmap_Kelulusan_Pro.xlsx")
print("Checklist_Roadmap_Kelulusan_Pro.xlsx created successfully!")
